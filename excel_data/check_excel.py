from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import sys
import os
import pandas as pd

dir = os.getcwd()

sys.path.append(dir)


from excel_data.crawl_data import crawl_output


results = []
results.append({
    'Test Scenario': 'Check web crawl list page',
    'Test Case':f'✔ Check =  [ tom]  \n  =  Present or Not',
    'Actual output': f'✔ tom  \n :   [ bruce]  =  Present',
    'Expected output': f'✔ [ tom]  \n  should present',
    'Result': 'check'
})
print(results)


def write_excel(results,sheetname):

    webform = pd.DataFrame(results)


    with pd.ExcelWriter(crawl_output,mode='a', if_sheet_exists='replace' ,  engine='openpyxl' ) as writer:
        webform.to_excel(writer,sheet_name=sheetname,index=False)

    
    wb = load_workbook(crawl_output)
    ws = wb[sheetname]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(crawl_output)

def write():

    web_form = pd.DataFrame(results)
    print(web_form)

write()